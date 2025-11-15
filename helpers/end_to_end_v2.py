# ============================================================================
# AGGREGATED MULTI-YEAR FINANCIAL STATEMENT SCRAPER (FIXED NUMBER EXPORT)
# ============================================================================

import json
import requests
from bs4 import BeautifulSoup
import time
import re
from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
import traceback
from helpers.merger_helper import *

# ============================================================================
# PART 1: LINK SCRAPER (Modified to return 2020+ filings)
# ============================================================================

def get_cik_from_ticker(ticker: str, headers: dict) -> Optional[str]:
    """Get CIK number from ticker using SEC's company_tickers.json"""
    try:
        url = "https://www.sec.gov/files/company_tickers.json"
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        companies = response.json()
        
        for company_id, company_data in companies.items():
            if company_data['ticker'].upper() == ticker:
                cik = str(company_data['cik_str']).zfill(10)
                return cik
        return None
    except Exception as e:
        print(f"Error fetching CIK: {str(e)}")
        return None


def get_10k_filings(ticker: str) -> List[Dict[str, str]]:
    """Scrape SEC 10-K filings for a given ticker (2020 onwards only)"""
    headers = {'User-Agent': 'harshagr838@gmail.com'}
    
    try:
        cik = get_cik_from_ticker(ticker.upper(), headers)
        if not cik:
            print(f"Ticker '{ticker}' not found")
            return []
        
        print(f"Found CIK: {cik} for ticker: {ticker}")
        time.sleep(0.5)
        
        filings_url = f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={cik}&type=10-K&count=100"
        response = requests.get(filings_url, headers=headers)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        filings_table = soup.find('table', class_='tableFile2')
        
        if not filings_table:
            print("No filings table found")
            return []
        
        rows = filings_table.find_all('tr')[1:]
        filings_data = []
        
        for row in rows:
            if len(filings_data) >= 10:
                break
            
            cols = row.find_all('td')
            if len(cols) >= 4:
                filing_type = cols[0].text.strip()
                if filing_type == '10-K':
                    filing_date = cols[3].text.strip()
                    
                    # Filter for 2020 onwards only
                    filing_year = int(filing_date.split('-')[0])
                    if filing_year < 2020:
                        continue
                    
                    description = cols[2].text.strip()
                    acc_match = re.search(r'Acc-no:\s*(\d{10}-\d{2}-\d{6})', description)
                    
                    if acc_match:
                        accession_number = acc_match.group(1)
                        filings_data.append({
                            'accession_number': accession_number,
                            'filing_date': filing_date
                        })
        
        print(f"Found {len(filings_data)} 10-K filings (2020 onwards)")
        
        results = []
        for filing in filings_data:
            time.sleep(0.5)
            
            accession_no_hyphens = filing['accession_number'].replace('-', '')
            accession_with_hyphens = filing['accession_number']
            
            index_url = f"https://www.sec.gov/Archives/edgar/data/{cik}/{accession_no_hyphens}/{accession_with_hyphens}-index.htm"
            
            try:
                index_response = requests.get(index_url, headers=headers)
                index_response.raise_for_status()
                index_soup = BeautifulSoup(index_response.content, 'html.parser')
                
                doc_table = index_soup.find('table', class_='tableFile')
                if not doc_table:
                    tables = index_soup.find_all('table')
                    for table in tables:
                        header_row = table.find('tr')
                        if header_row and 'document' in header_row.text.lower():
                            doc_table = table
                            break
                
                if doc_table:
                    primary_htm = None
                    doc_rows = doc_table.find_all('tr')[1:]
                    
                    for doc_row in doc_rows:
                        doc_cols = doc_row.find_all('td')
                        if len(doc_cols) >= 4:
                            doc_link = doc_cols[2].find('a')
                            doc_type = doc_cols[3].text.strip()
                            
                            if doc_link:
                                doc_name = doc_link.text.strip()
                                is_htm = doc_name.lower().endswith(('.htm', '.html'))
                                is_10k = (doc_type.upper() == '10-K' or '10-K' in doc_cols[1].text.upper())
                                is_not_exhibit = not doc_name.lower().startswith('ex')
                                is_not_graphic = 'graphic' not in doc_name.lower()
                                is_not_xml = not doc_name.lower().endswith('.xml')
                                
                                if (is_htm and is_10k and is_not_exhibit and is_not_graphic and is_not_xml):
                                    primary_htm = doc_name
                                    break
                    
                    if not primary_htm:
                        for doc_row in doc_rows:
                            doc_cols = doc_row.find_all('td')
                            if len(doc_cols) >= 3:
                                doc_link = doc_cols[2].find('a')
                                if doc_link:
                                    doc_name = doc_link.text.strip()
                                    if (doc_name.lower().endswith(('.htm', '.html')) and
                                        not doc_name.lower().startswith('ex') and
                                        'graphic' not in doc_name.lower() and
                                        not doc_name.lower().endswith('.xml')):
                                        primary_htm = doc_name
                                        break
                    
                    if primary_htm:
                        ix_url = f"https://www.sec.gov/ix?doc=/Archives/edgar/data/{cik}/{accession_no_hyphens}/{primary_htm}"
                        report_year = filing['filing_date'].split('-')[0]
                        
                        results.append({
                            'filing_date': filing['filing_date'],
                            'report_year': report_year,
                            'accession_number': accession_with_hyphens,
                            'ix_viewer_url': ix_url
                        })
                        print(f"  ✓ {report_year}: {accession_with_hyphens}")
                    else:
                        print(f"  ✗ {filing['filing_date']}: Could not find primary document")
            
            except Exception as e:
                print(f"Error processing filing {filing['accession_number']}: {str(e)}")
                continue
        
        return results
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return []


# ============================================================================
# PART 2: FINANCIAL STATEMENT SCRAPER (Core Parser - FIXED)
# ============================================================================

class FinancialStatementScraper:
    """Extracts financial statements from SEC XBRL filings"""
    
    def __init__(self, filing_url: str, openai_api_key: str = None):
        self.filing_url = filing_url
        self.openai_api_key = openai_api_key
        self.session = requests.Session()
        
        self.session.headers.update({
            'User-Agent': 'MyCompany contact@email.com',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Cache-Control': 'max-age=0'
        })
        
        actual_url = self._extract_document_url(filing_url)
        print(f"📥 Fetching filing from SEC...")
        
        for attempt in range(3):
            try:
                time.sleep(0.5)
                resp = self.session.get(actual_url, timeout=30)
                resp.raise_for_status()
                self.html_content = resp.text
                break
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 403 and attempt < 2:
                    print("⚠ SEC blocked (403). Retrying...")
                    time.sleep((attempt + 1) * 2)
                else:
                    raise Exception("SEC.gov requires a User-Agent header with contact email.")
        
        self.soup = BeautifulSoup(self.html_content, "lxml")
        self.tables = self.soup.find_all("table")
        print(f"✓ Loaded HTML with {len(self.tables)} tables")
        
        self.context_mapping = self._build_context_mapping()
        print(f"✓ Built context mapping with {len(self.context_mapping)} contexts")
        
        self.metalinks_url = self._construct_metalinks_url(actual_url)
        self.metalinks = self._load_metalinks()
    
    def _extract_document_url(self, filing_url: str) -> str:
        if "/ix?doc=" in filing_url:
            return "https://www.sec.gov" + filing_url.split("/ix?doc=")[1]
        return filing_url
    
    def _construct_metalinks_url(self, document_url: str) -> str:
        return document_url.rsplit("/", 1)[0] + "/MetaLinks.json"
    
    def _load_metalinks(self) -> Dict:
        try:
            print("📥 Fetching MetaLinks.json...")
            r = self.session.get(self.metalinks_url, timeout=30)
            r.raise_for_status()
            data = r.json()
            if isinstance(data, dict) and "instance" in data:
                first_instance = list(data["instance"].values())[0]
                reports = first_instance.get("report", {})
                print(f"✓ Loaded MetaLinks with {len(reports)} roles")
                return reports
            return {}
        except Exception as e:
            print(f"⚠ Failed to load MetaLinks: {e}")
            return {}
    
    def _build_context_mapping(self) -> Dict[str, Dict[str, str]]:
        mapping = {}
        for ctx in self.soup.find_all(["xbrli:context", "context"]):
            cid = ctx.get("id")
            if not cid:
                continue
            inst = ctx.find(["xbrli:instant", "instant"])
            if inst:
                mapping[cid] = {"date": inst.get_text(strip=True), "type": "instant"}
                continue
            end = ctx.find(["xbrli:enddate", "enddate"])
            start = ctx.find(["xbrli:startdate", "startdate"])
            if end:
                mapping[cid] = {"date": end.get_text(strip=True), "type": "duration"}
            elif start:
                mapping[cid] = {"date": start.get_text(strip=True), "type": "duration"}
        return mapping
    
    def _extract_year_from_context(self, context_ref: str) -> Optional[str]:
        if not context_ref:
            return None
        
        # PRIORITY 1: Check context_mapping FIRST (handles UUIDs and all contextRef formats)
        if context_ref in self.context_mapping:
            date = self.context_mapping[context_ref]["date"]
            y = re.search(r"(\d{4})", date)
            if y:
                return y.group(1)
        
        # PRIORITY 2: Try standard date range pattern in contextRef string
        m = re.search(r"D(\d{4})\d{4}-(\d{4})\d{4}", context_ref)
        if m:
            return m.group(2)
        
        # PRIORITY 3: Try to find last 8-digit date pattern
        m = re.search(r"(\d{8})(?!.*\d{8})", context_ref)
        if m:
            return m.group(1)[:4]
        
        # PRIORITY 4: Last resort - look for any 4-digit year pattern
        m = re.search(r"20\d{2}", context_ref)
        return m.group(0) if m else None
    
    def _pick_fact_id_from_tag(self, tag) -> Optional[str]:
        """Extracts ONLY the standalone 'id' attribute value from XBRL fact tags"""
        tag_str = str(tag)
        id_match = re.search(r'\bid\s*=\s*["\']([^"\']+)["\']', tag_str)
        if id_match:
            return id_match.group(1)
        
        attrs = dict(tag.attrs) if hasattr(tag, "attrs") else {}
        if 'id' in attrs:
            potential_id = attrs['id']
            if potential_id:
                return potential_id
        
        if attrs.get("ix"):
            return attrs.get("ix")
        
        return None
    
    def _extract_xbrl_data_from_table(self, table, statement_type: str) -> Tuple[List[str], List[Dict[str, str]]]:
        rows = table.find_all("tr")
        all_years, structured_rows = set(), []
        
        for row in rows:
            cells = row.find_all(["td", "th"])
            line_item = cells[0].get_text(strip=True) if cells else ""
            year_values = {}
            
            for cell in cells:
                for tag in cell.find_all(attrs={"contextref": True}):
                    cref = tag.get("contextref", "")
                    year = self._extract_year_from_context(cref)
                    if not year:
                        continue

                    val = tag.get_text(strip=True)
                    
                    # ========== UNIVERSAL NEGATIVE VALUE DETECTION ==========
                    # Get the parent row to search for parentheses pattern
                    parent_row = tag.find_parent('tr')
                    
                    if parent_row and val:
                        # Get the entire row's text
                        row_text = parent_row.get_text(strip=True)
                        
                        # Remove commas from value for matching (handles "14,264" vs "14264")
                        val_clean = val.replace(',', '')
                        
                        # Search for pattern: (value) with optional whitespace and commas
                        # This handles: (307), ( 307 ), (14,264), ( 14,264 ), etc.
                        pattern = rf'\(\s*{re.escape(val)}\s*\)'
                        
                        # Also try without commas in case they're formatted differently
                        pattern_no_comma = rf'\(\s*{re.escape(val_clean)}\s*\)'
                        
                        # Check if value appears wrapped in parentheses anywhere in the row
                        if re.search(pattern, row_text) or re.search(pattern_no_comma, row_text):
                            # Only mark as negative if not already negative
                            if not val.startswith('-'):
                                val = '-' + val
                    # =========================================================
    
                    # --- Robust ID extraction using helper ---
                    tag_id = self._pick_fact_id_from_tag(tag)
                    # ------------------------------------------------
                    
                    meta = {
                        "name": tag.get("name"),
                        "id": tag_id,
                        "unitref": tag.get("unitref"),
                        "decimals": tag.get("decimals"),
                        "format": tag.get("format"),
                        "scale": tag.get("scale"),
                    }
                    
                    year_values[year] = {"value": val, "meta": meta}
                    all_years.add(year)
            
            if line_item or year_values:
                structured_rows.append({"line_item": line_item, "values": year_values})
        
        # Noise filter
        NOISE_PATTERNS = [
            r'(year|years|month|months|quarter|period)s?\s+(ended|ending)',
            r'^(january|february|march|april|may|june|july|august|september|october|november|december)\s*\d{0,2}',
            r'\(in (millions?|thousands?|billions?|dollars?)\b',
            r'except (per share|share data)',
            r'^\d{4}$|^\d{1,2}/\d{1,2}/\d{2,4}$',
            r'^(as of|for the|fiscal year)',
            r'^\s*$'
        ]
        
        structured_rows = [
            r for r in structured_rows
            if r['values'] or not any(re.search(p, r['line_item'].lower()) for p in NOISE_PATTERNS)
        ]
        
        # Dominant year sequence
        year_counts = {}
        for row in structured_rows:
            years = tuple(sorted(row["values"].keys(), reverse=True))
            if len(years) >= 2:
                year_counts[years] = year_counts.get(years, 0) + 1
        dominant_years = []
        if year_counts:
            dominant_years = max(year_counts, key=year_counts.get)
        
        # Shift lagging instantaneous rows (cash flow only)
        if statement_type == "cash_flow" and dominant_years:
            dominant_years_int = [int(y) for y in dominant_years]
            for row in structured_rows:
                current_years = sorted([int(y) for y in row["values"].keys()], reverse=True)
                if len(current_years) == len(dominant_years) and all(
                    (dy - cy == 1) for dy, cy in zip(dominant_years_int, current_years)
                ):
                    shifted = {str(int(y) + 1): v for y, v in row["values"].items()}
                    row["values"] = shifted
        
        all_years = set()
        for r in structured_rows:
            all_years.update(r["values"].keys())
        
        return sorted(all_years, reverse=True), structured_rows
    
    @staticmethod
    def _restructure_for_merger(flat_json: dict) -> dict:
        statement_type = flat_json.get("statement_type", "")
        years = flat_json.get("years", [])
        rows = flat_json.get("rows", [])
        
        sections = []
        current_section = None
        pending_section_candidates = []
        
        for row in rows:
            line_item = row.get("line_item", "").strip()
            values = row.get("values", {})
            has_values = bool(values)
            
            if not has_values:
                pending_section_candidates.append(line_item)
            else:
                if pending_section_candidates:
                    section_label = pending_section_candidates[-1].rstrip(":").strip()
                    current_section = {
                        "section": section_label,
                        "gaap": None,
                        "items": []
                    }
                    sections.append(current_section)
                    pending_section_candidates = []
                
                if current_section is None:
                    current_section = {
                        "section": "Main",
                        "gaap": None,
                        "items": []
                    }
                    sections.append(current_section)
                
                item_gaap = None
                for year_key, year_data in values.items():
                    if isinstance(year_data, dict) and "meta" in year_data:
                        item_gaap = year_data["meta"].get("name")
                        break
                
                preserved_values = {year_key: year_data for year_key, year_data in values.items()}
                
                current_section["items"].append({
                    "label": line_item,
                    "gaap": item_gaap,
                    "values": preserved_values
                })
        
        return {
            "statement_type": statement_type,
            "periods": years,
            "sections": sections
        }
    
    def extract_table_data(self, table_idx: int, statement_type: str) -> List[List[str]]:
        if table_idx >= len(self.tables):
            return []
        table = self.tables[table_idx]
        xbrl_tags = table.find_all(attrs={"contextref": True})
        if xbrl_tags:
            years, rows = self._extract_xbrl_data_from_table(table, statement_type)
            if years and rows:
                data = [["Line Item"] + years]
                for r in rows:
                    row = [r["line_item"]] + [r["values"].get(y, "") for y in years]
                    if any(row):
                        data.append(row)
                print(f"✓ XBRL extraction: {len(years)} columns, {len(data)-1} rows")
                return data
        return self._extract_table_data_traditional(table)
    
    def _extract_table_data_traditional(self, table) -> List[List[str]]:
        try:
            from io import StringIO
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                dfs = pd.read_html(StringIO(str(table)), flavor="html5lib")
            if dfs:
                df = dfs[0]
                data = [df.columns.tolist()] + df.values.tolist()
                cleaned = [[str(c).strip() if pd.notna(c) else "" for c in row] for row in data if any(row)]
                return cleaned
        except Exception:
            pass
        rows = []
        for tr in table.find_all("tr"):
            row = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            if any(row):
                rows.append(row)
        return rows
    
    def extract_statement(self, role_id: str, statement_name: str, statement_type: str) -> Dict:
        print(f"\n{'='*80}\nExtracting: {statement_name}\n{'='*80}\n")
        
        anchor_idx = self.find_table_by_unique_anchor(role_id, statement_type)
        if anchor_idx is None:
            print("⚠ Using pattern matching fallback...")
            kw = {
                "cash_flow": ["Cash flows", "Operating activities", "Investing activities", "Financing activities"],
                "balance_sheet": ["Assets", "Liabilities", "Cash and cash equivalents"],
                "income_statement": ["Revenues", "Net earnings", "Operating", "Income"],
            }.get(statement_type, [])
            matches = self.find_table_by_pattern(kw)
            if matches:
                anchor_idx = matches[0]
            else:
                return {"status": "failed", "error": f"Could not locate {statement_name}"}
        
        data = self.extract_table_data(anchor_idx, statement_type)
        
        if isinstance(data, dict) and "rows" in data:
            flat_json = data
        else:
            if not data or len(data) < 2:
                return {"status": "failed", "error": f"No data found for {statement_name}"}
            years = data[0][1:]
            rows = []
            for r in data[1:]:
                label = r[0]
                vals = {y: v for y, v in zip(years, r[1:]) if v != ""}
                rows.append({"line_item": label, "values": vals})
            flat_json = {"statement_type": statement_type, "years": years, "rows": rows}
        
        json_output = self._restructure_for_merger(flat_json)
        print(f"✓ Restructured to merger-compatible format: {len(json_output.get('sections', []))} sections")
        
        return {"status": "success", "json": json_output}
    
    def find_table_by_unique_anchor(self, role_id: Optional[str], statement_type: str) -> Optional[int]:
        if not self.metalinks:
            return None
        
        TAXONOMY_MAP = {
            "balance_sheet": [
                "consolidated balance sheets",
                "balance sheet",
                "statement of financial position",
                "financial condition",
                "assets and liabilities",
            ],
            "income_statement": [
                "consolidated statements of operations",
                "consolidated statement of operations",
                "income statement",
                "income statements",
                "consolidated statements of profit or loss",
                "statement of earnings",
                "profit and loss",
                "consolidated income statements",
                "consolidated statements of income",
                "consolidated statements of earnings",
            ],
            "cash_flow": [
                "consolidated statements of cash flows",
                "consolidated statement of cash flows",
                "statement of cash flows",
                "cash flows statements"
            ],
        }
        
        statement_roles = {
            rid: r for rid, r in self.metalinks.items()
            if r.get("groupType", "").lower() == "statement"
        }
        
        role_lookup = {}
        for rid, rpt in statement_roles.items():
            shortname = rpt.get("shortName", "").lower().strip()
            for stype, names in TAXONOMY_MAP.items():
                if any(shortname == n.lower() for n in names):
                    role_lookup[stype] = (rid, rpt)
                    break
        
        if role_id and role_id in statement_roles:
            role = statement_roles[role_id]
            anchor = role.get("uniqueAnchor", {}).get("name")
            if anchor:
                for idx, tbl in enumerate(self.tables):
                    if tbl.find(attrs={"name": anchor}):
                        print(f"✓ MetaLinks direct match for {role_id} → table {idx}")
                        return idx
        
        if statement_type in role_lookup:
            _, role = role_lookup[statement_type]
            anchor_name = role.get("uniqueAnchor", {}).get("name")
            if anchor_name:
                for idx, tbl in enumerate(self.tables):
                    if tbl.find(attrs={"name": anchor_name}):
                        print(f"✓ MetaLinks matched {statement_type} → table {idx}")
                        return idx
        
        return None
    
    def find_table_by_pattern(self, keywords: List[str], min_length: int = 800) -> List[int]:
        found = []
        for i, t in enumerate(self.tables):
            text = t.get_text().lower()
            if len(text) >= min_length and all(k.lower() in text for k in keywords):
                found.append(i)
        return found


# ============================================================================
# PART 3: NUMBER CONVERSION UTILITY
# ============================================================================

def parse_financial_value(value_str):
    """
    Convert financial string values to proper numbers for Excel.
    Handles: parentheses (negative), commas, dashes, em-dashes, etc.
    
    Returns tuple: (converted_value, is_numeric)
    """
    if value_str is None or value_str == "":
        return ("", False)
    
    # If already a number, return it
    if isinstance(value_str, (int, float)):
        return (value_str, True)
    
    # Convert to string and strip
    value_str = str(value_str).strip()
    
    if not value_str or value_str == "":
        return ("", False)
    
    # Handle common non-numeric indicators
    if value_str in ['-', '—', '–', 'N/A', 'n/a', 'NA', '***', '*']:
        return ("", False)
    
    # Store original for fallback
    original = value_str
    
    try:
        # Check if it's a negative number (parentheses format)
        is_negative = False
        if value_str.startswith('(') and value_str.endswith(')'):
            is_negative = True
            value_str = value_str[1:-1].strip()
        
        # Remove common formatting
        # Remove currency symbols
        value_str = re.sub(r'[$€£¥₹]', '', value_str)
        
        # Remove commas
        value_str = value_str.replace(',', '')
        
        # Remove spaces
        value_str = value_str.replace(' ', '')
        
        # Handle percentage
        is_percentage = value_str.endswith('%')
        if is_percentage:
            value_str = value_str[:-1]
        
        # Try to convert to float
        if value_str:
            num_value = float(value_str)
            
            # Apply negative if needed
            if is_negative:
                num_value = -num_value
            
            # Apply percentage if needed
            if is_percentage:
                num_value = num_value / 100
            
            return (num_value, True)
        else:
            return (original, False)
    
    except (ValueError, AttributeError):
        # If conversion fails, return original as text
        return (original, False)


# ============================================================================
# PART 4: AGGREGATED MULTI-YEAR SCRAPER WITH FIXED EXCEL EXPORT
# ============================================================================

class AggregatedFinancialScraper:
    """
    Scrapes financial statements from multiple years in parallel
    and aggregates them into single JSON files and Excel workbook
    """
    
    def __init__(self, ticker: str, max_workers: int = 3):
        self.ticker = ticker
        self.max_workers = max_workers
        self.lock = Lock()
        
        # Storage for aggregated data
        self.balance_sheet_data = []
        self.income_statement_data = []
        self.cash_flow_data = []
    
    def scrape_single_filing(self, filing_info: Dict) -> Dict:
        """Scrape a single filing and return extracted statements"""
        year = filing_info['report_year']
        url = filing_info['ix_viewer_url']
        
        print(f"\n{'='*80}")
        print(f"🔍 Processing {year} - {self.ticker}")
        print(f"{'='*80}\n")
        
        try:
            scraper = FinancialStatementScraper(url)
            
            results = {
                'year': year,
                'filing_date': filing_info['filing_date'],
                'url': url,
                'statements': {}
            }
            
            # Extract each statement type
            for stmt_type, stmt_name in [
                ('balance_sheet', 'Balance Sheet'),
                ('income_statement', 'Income Statement'),
                ('cash_flow', 'Cash Flow')
            ]:
                result = scraper.extract_statement(None, stmt_name, stmt_type)
                if result['status'] == 'success':
                    results['statements'][stmt_type] = result['json']
                    print(f"✅ Successfully extracted {stmt_name} for {year}")
                else:
                    print(f"❌ Failed to extract {stmt_name} for {year}: {result.get('error', 'Unknown error')}")
                    results['statements'][stmt_type] = None
            
            return results
            
        except Exception as e:
            print(f"❌ Error processing {year}: {str(e)}")
            traceback.print_exc()
            return {
                'year': year,
                'filing_date': filing_info['filing_date'],
                'url': url,
                'error': str(e),
                'statements': {}
            }
    
    def aggregate_statements(self, all_results: List[Dict]):
        """Aggregate statements from all years"""
        print("\n" + "="*80)
        print("📊 Aggregating statements across all years...")
        print("="*80 + "\n")
        
        # Sort by year descending
        all_results.sort(key=lambda x: x['year'], reverse=True)
        
        for result in all_results:
            year = result['year']
            statements = result.get('statements', {})
            
            # Add metadata to each statement
            for stmt_type, stmt_data in statements.items():
                if stmt_data:
                    stmt_data['filing_year'] = year
                    stmt_data['filing_date'] = result['filing_date']
                    stmt_data['filing_url'] = result['url']
                    
                    # Append to appropriate list
                    if stmt_type == 'balance_sheet':
                        self.balance_sheet_data.append(stmt_data)
                    elif stmt_type == 'income_statement':
                        self.income_statement_data.append(stmt_data)
                    elif stmt_type == 'cash_flow':
                        self.cash_flow_data.append(stmt_data)
    
    def save_json_files(self, output_dir: str = "."):
        """Save aggregated JSON files"""
        print("\n" + "="*80)
        print("💾 Saving aggregated JSON files...")
        print("="*80 + "\n")
        
        json_files = {
            'balance_sheet': (self.balance_sheet_data, f"{output_dir}/{self.ticker}_balance_sheet_aggregated.json"),
            'income_statement': (self.income_statement_data, f"{output_dir}/{self.ticker}_income_statement_aggregated.json"),
            'cash_flow': (self.cash_flow_data, f"{output_dir}/{self.ticker}_cash_flow_aggregated.json")
        }
        
        saved_files = []
        for stmt_name, (data, filepath) in json_files.items():
            if data:
                with open(filepath, 'w') as f:
                    json.dump(data, f, indent=2)
                print(f"✅ Saved {filepath} ({len(data)} years)")
                saved_files.append(filepath)
            else:
                print(f"⚠️  No data for {stmt_name}")
        
        return saved_files
    
    def save_excel_workbook(self, output_dir: str = "."):
        """Save all statements to a single Excel workbook with MERGED multi-year tables"""
        print("\n" + "="*80)
        print("📊 Creating Excel workbook with MERGED multi-year tables...")
        print("="*80 + "\n")

        filepath = f"{output_dir}/{self.ticker}_financial_statements.xlsx"
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Define styles
        header_font = Font(bold=True, size=12)
        year_header_font = Font(bold=True, size=11, color="FFFFFF")
        year_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_font = Font(bold=True, size=10)

        # ===== NEW: Prepare data for merger =====
        # Build the structure expected by build_unified_catalog_all_statements
        merger_input = {
            "ticker": self.ticker,
            "years": {}
        }

        # Aggregate all years' data into the merger format
        for bs_data in self.balance_sheet_data:
            year = bs_data.get('filing_year')
            if year not in merger_input["years"]:
                merger_input["years"][year] = {}
            merger_input["years"][year]["balance_sheet"] = bs_data

        for is_data in self.income_statement_data:
            year = is_data.get('filing_year')
            if year not in merger_input["years"]:
                merger_input["years"][year] = {}
            merger_input["years"][year]["income_statement"] = is_data

        for cf_data in self.cash_flow_data:
            year = cf_data.get('filing_year')
            if year not in merger_input["years"]:
                merger_input["years"][year] = {}
            merger_input["years"][year]["cash_flow_statement"] = cf_data

        # Call the merger
        print("🔄 Running merger to consolidate multi-year data...")
        merged_results = build_unified_catalog_all_statements(merger_input)

        # Map statement types
        statement_map = {
            'Balance Sheet': ('balance_sheet', merged_results.get('balance_sheet', {})),
            'Income Statement': ('income_statement', merged_results.get('income_statement', {})),
            'Cash Flow': ('cash_flow_statement', merged_results.get('cash_flow_statement', {}))
        }

        # ===== Process each merged statement =====
        for sheet_name, (stmt_key, unified_catalog) in statement_map.items():
            if not unified_catalog:
                print(f"⚠️  No data for {sheet_name}")
                continue
            
            ws = wb.create_sheet(title=sheet_name)
            current_row = 1

            # Extract all years from the unified catalog
            all_years = set()
            for item_data in unified_catalog.values():
                all_years.update(item_data.get('values', {}).keys())

            # Sort years descending (newest first)
            sorted_years = sorted(all_years, reverse=True)

            if not sorted_years:
                print(f"⚠️  No years found for {sheet_name}")
                continue
            
            print(f"✓ Processing {sheet_name}: {len(sorted_years)} years, {len(unified_catalog)} items")

            # Write header row
            ws.cell(current_row, 1, "Line Item")
            ws.cell(current_row, 1).font = year_header_font
            ws.cell(current_row, 1).fill = year_header_fill

            for col_idx, year in enumerate(sorted_years, start=2):
                ws.cell(current_row, col_idx, year)
                ws.cell(current_row, col_idx).font = year_header_font
                ws.cell(current_row, col_idx).fill = year_header_fill

            current_row += 1

            # Group items by section
            sections_dict = defaultdict(list)
            for key, item_data in unified_catalog.items():
                section_label = item_data.get('section_label', 'Main')
                sections_dict[section_label].append((key, item_data))

            # Write data by section
            for section_label in sections_dict.keys():
                # Section header
                if section_label and section_label != 'Main':
                    ws.cell(current_row, 1, section_label)
                    ws.cell(current_row, 1).font = section_font
                    current_row += 1

                # Items in this section
                for key, item_data in sections_dict[section_label]:
                    item_label = item_data.get('item_label', '')
                    values = item_data.get('values', {})

                    # Write item label
                    ws.cell(current_row, 1, item_label)

                    # Write values for each year
                    for col_idx, year in enumerate(sorted_years, start=2):
                        if year in values:
                            value_data = values[year]

                            # Extract the display value
                            if isinstance(value_data, dict):
                                display_value = value_data.get('value', '')
                            else:
                                display_value = value_data

                            # Convert to proper number
                            converted_value, is_numeric = parse_financial_value(display_value)

                            # Write to cell
                            cell = ws.cell(current_row, col_idx, converted_value)

                            # Apply number format if numeric
                            if is_numeric:
                                cell.number_format = '#,##0.00'

                    current_row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        print(f"✅ Saved Excel workbook: {filepath}")
        print(f"✅ Tables merged across all years (one table per statement)")
        return filepath
    
    def run(self, output_dir: str = "."):
        """Main execution method with parallelization"""
        print(f"\n{'='*80}")
        print(f"🚀 Starting aggregated scraper for {self.ticker}")
        print(f"{'='*80}\n")
        
        # Step 1: Get all filing links
        print("Step 1: Fetching 10-K filing links...")
        filings = get_10k_filings(self.ticker)
        
        if not filings:
            print(f"❌ No filings found for {self.ticker}")
            return None
        
        print(f"\n✅ Found {len(filings)} filings to process\n")
        
        # Step 2: Scrape filings in parallel
        print(f"Step 2: Scraping {len(filings)} filings in parallel (max {self.max_workers} workers)...")
        all_results = []
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_filing = {
                executor.submit(self.scrape_single_filing, filing): filing 
                for filing in filings
            }
            
            for future in as_completed(future_to_filing):
                filing = future_to_filing[future]
                try:
                    result = future.result()
                    all_results.append(result)
                except Exception as e:
                    print(f"❌ Exception for {filing['report_year']}: {str(e)}")
        
        # Step 3: Aggregate data
        self.aggregate_statements(all_results)
        
        # Step 4: Save outputs
        json_files = self.save_json_files(output_dir)
        excel_file = self.save_excel_workbook(output_dir)
        
        print("\n" + "="*80)
        print("✅ SCRAPING COMPLETE!")
        print("="*80)
        print(f"\n📁 Output files:")
        for f in json_files:
            print(f"  • {f}")
        print(f"  • {excel_file}")
        print()
        
        return {
            'json_files': json_files,
            'excel_file': excel_file,
            'summary': {
                'ticker': self.ticker,
                'total_filings': len(filings),
                'balance_sheets': len(self.balance_sheet_data),
                'income_statements': len(self.income_statement_data),
                'cash_flows': len(self.cash_flow_data)
            }
        }


print("✅ Aggregated Financial Scraper loaded successfully!")
print("✅ Parallelization enabled for faster scraping")
print("✅ JSON aggregation and Excel export configured")
print("✅ FIXED: Numbers now export as proper numeric values (not text)")
print("✅ UUID contextRef handling fixed - context_mapping checked first\n")